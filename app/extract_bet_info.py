import easyocr
import glob, os
import re
import requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill

reader = easyocr.Reader(['en'])
folder = os.getcwd()
league = ""
resp = ""

# teams dictionary
premierLeague = {
                "Arsenal" : "Arsenal",
                "Chelsea" : "Chelsea",
                "Liverpool" : "Liverpool", 
                "Manchester City" : "Manchester City",
                "Manchester United" : "Manchester United"
                }

laLiga = {
        "Atletico Madrid" : "Atlético Madrid", 
        "Barcelona" : "Barcelona",
        "Real Madrid" : "Real Madrid"
        }

bundesLiga = {
            "Bayer Leverkusen" : "Bayer Leverkusen",
            "Bayern Munchen" : "Bayern Munich", 
            "Borussia Dortmund" : "Borussia Dortmund", 
            "Eintracht Frankfurt" : "Eintracht Frankfurt",
            "Mainz" : "Mainz",
            "RB Leipzig" : "RB Leipzig",
            "Stuttgart" : "Stuttgart"
            }

serieA = {
    "AC Milan" : "AC Milan", 
    "Inter Milan" : "Inter Milan",
    "Juventus" : "Juventus",
    "Lazio" : "Lazio",
    "Napoli" : "Napoli",
    "AS Roma" : "Roma"
        }   

ligue1 = {
        "Lille" : "Lille",
        "Lyon" : "Lyon",
        "Olympique Marseille" : "Marseille",
        "Monaco" : "Monaco",
        "Nice" : "Nice",
        "PSG" : "Paris SG"
        }

leagues ={
        "England" : premierLeague,
        "Spain" : laLiga,
        "Germany" : bundesLiga,
        "Italy" : serieA,
        "France" : ligue1
        }

teams_keys = []
for d in [premierLeague, laLiga, bundesLiga, serieA, ligue1]:
    for key in d:
        teams_keys.append(key)
teams_values = list(premierLeague.values()) + list(laLiga.values()) + list(bundesLiga.values()) + list(serieA.values()) + list(ligue1.values())
played_matches = {key: 0 for key in teams_keys}
won_matches = {key:0 for key in teams_keys}

# generate new Excel sheet
file_path = "placed_bets.xlsx"
sheet_name = "Bets"

if os.path.exists(file_path):
    wb = load_workbook(file_path)
    sheet=wb[sheet_name]
    total_wins = int(sheet["L2"].value)
    total_losses = int(sheet["M2"].value)
    total_won = round(float(sheet["P2"].value), 2)
    invested = int(sheet["O2"].value)
    sum_odds = round(float(sheet["S2"].value * (total_losses + total_wins)), 2)
    balance = round(float(sheet["T2"].value), 2)
else:
    wb = Workbook()
    wb.create_sheet("Bets")
    wb.create_sheet("Teams")
    wb.remove(wb["Sheet"])
    sheet=wb[sheet_name]
    total_wins = 0
    total_losses = 0
    total_won = 0
    invested = 0
    sum_odds = 0
    balance = 1000

    sheet["A1"] = "Date"
    sheet["B1"] = "Home"
    sheet["C1"] = "Away"
    sheet["D1"] = "Bet"
    sheet["E1"] = "Odds"
    sheet["F1"] = "Wager"
    sheet["G1"] = "Win"
    sheet["H1"] = "Profit"
    sheet["I1"] = "Result"
    sheet["J1"] = "Success"

    sheet["L1"] = "Wins"
    sheet["M1"] = "Losses"
    sheet["N1"] = "Win percentage"
    sheet["O1"] = "Invested"
    sheet["P1"] = "Won"
    sheet["Q1"] = "Gained"
    sheet["R1"] = "Return"
    sheet["S1"] = "Average odds"
    sheet["T1"] = "Balance"

wb.save(file_path)

# function to get match results from TheSportsDB API
def retrieve_data(index, img_path):
    global balance, invested, league, sum_odds, total_wins, total_losses, total_won, resp
    result = reader.readtext(img_path)

    content = []
    for detection in result:
        content.append(detection[1])

    filename = os.path.splitext(os.path.basename(img_path))[0]
    filename = filename.split('.')[0]
    parts = re.split(r'[_\W]+', filename)
    if len(parts) >= 3:
        year, month, day = parts[0:3]
        month = month.zfill(2)
        day = day.zfill(2)
        date_obj = datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
        date = date_obj.strftime("%d.%m.%Y")
        formated_date = date_obj.strftime("%Y-%m-%d")

    home = content[0]
    away = content[1]
    bet = content[3]
    odds = content[4]
    odds = float(odds)
    sum_odds += odds
    wager = content[8]
    wager = float(wager.replace("RON", "").strip().replace(",", "."))
    invested += wager
    win = odds * wager
    profit = win - wager

    sheet[f"A{index}"] = date
    sheet[f"B{index}"] = home
    sheet[f"C{index}"] = away
    sheet[f"D{index}"] = bet
    sheet[f"E{index}"] = odds
    sheet[f"F{index}"] = wager
    sheet[f"G{index}"] = win
    sheet[f"H{index}"] = profit

    if home in ligue1 or away in ligue1:
        url = f"https://www.thesportsdb.com/api/v1/json/123/eventsday.php?d={formated_date}&l=French%20Ligue%201"
        resp = requests.get(url)
        data = resp.json()
    else:
        if home in premierLeague or away in premierLeague:
            league = "English_Premier_League"
            
        if home in laLiga or away in laLiga:
            league = "Spanish_La_Liga"

        if home in bundesLiga or away in bundesLiga:
            league = "German_Bundesliga"
            
        if home in serieA or away in serieA:
            league = "Italian_Serie_A"

        url = f"https://www.thesportsdb.com/api/v1/json/123/eventsday.php?d={formated_date}&l={league}"
        resp = requests.get(url)
        data = resp.json()

    home_goals = ""
    away_goals = ""
    for event in data.get("events", []):
        if home in teams_keys:
            corresp = teams_keys.index(home)
            home = teams_values[corresp]
        if away in teams_keys:
            corresp = teams_keys.index(away)
            away = teams_values[corresp]
			
        if home in event["strEvent"] or away in event["strEvent"]:
            home_goals = event.get("intHomeScore")
            away_goals = event.get("intAwayScore")
            print(f"Result: {home} {home_goals} - {away_goals} {away} on {event['dateEvent']}")
            
    home_goals = int(home_goals)
    away_goals = int(away_goals)
    sheet[f"I{index}"] = str(home_goals) + " - " + str(away_goals)

    if bet == "1":
        if home_goals > away_goals:
            sheet[f"J{index}"].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            sheet[f"J{index}"] = 1
            total_wins += 1
            total_won += win
            balance += profit
        else:  
            sheet[f"J{index}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            sheet[f"J{index}"] = 0
            total_losses += 1
            balance -= wager

    if bet == "2":
        if home_goals < away_goals:
            sheet[f"J{index}"].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            sheet[f"J{index}"] = 1
            total_wins += 1
            total_won += win
            balance += profit

        else:  
            sheet[f"J{index}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            sheet[f"J{index}"] = 0
            total_losses += 1
            balance -= wager

    if bet == "1X" or bet == "1x" or bet == "lX" or bet == "lx" or bet == "IX" or bet == "Ix":
        if home_goals >= away_goals:
            sheet[f"J{index}"].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            sheet[f"J{index}"] = 1
            total_wins += 1
            total_won += win
            balance += profit

        else:  
            sheet[f"J{index}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            sheet[f"J{index}"] = 0
            total_losses += 1
            balance -= wager

    if bet == "X2" or bet == "x2":
        if home_goals <= away_goals:
            sheet[f"J{index}"].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            sheet[f"J{index}"] = 1
            total_wins += 1
            total_won += win
            balance += profit

        else:  
            sheet[f"J{index}"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            sheet[f"J{index}"] = 0
            total_losses += 1
            balance -= wager
    
    wb.save(file_path)

# call function on current batch of images
index = sheet.max_row

for filename in os.listdir(folder):
    if filename.lower().endswith(".jpg"):
        img_path = os.path.join(folder, filename)
        retrieve_data(index + 1, img_path)
        index += 1

sheet["L2"] = total_wins
sheet["M2"] = total_losses
win_perc = round(total_wins / (total_wins + total_losses) * 100, 2)
sheet["N2"] = win_perc
sheet["O2"] = invested
sheet["P2"] = total_won
sheet["Q2"] = total_won - invested
sheet["R2"] = str(round(balance/1000 -1, 2)) + "%"
avg_odds = round(float(sum_odds / (index - 1)), 2)
sheet["S2"] = avg_odds
sheet["T2"] = balance

wb.save(file_path)

# document styling
sheet["L1"].fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
sheet["M1"].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
sheet["N1"].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
sheet["O1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
sheet["P1"].fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
sheet["Q1"].fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
sheet["R1"].fill = PatternFill(start_color="8497B0", end_color="8497B0", fill_type="solid")
sheet["S1"].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
sheet["T1"].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

for row in range(1,sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row,col).font = Font(name='Helvetica', size=12, bold=True, color = '000000')
        sheet.cell(row,col).alignment = Alignment(horizontal='center', vertical='center')

for col in sheet.columns:
     max_width = 0
     column = col[0].column_letter
     for cell in col:
             if len(str(cell.value)) > max_width:
                 max_width = len(str(cell.value))
     set_col_width = max_width + 7
     sheet.column_dimensions[column].width = set_col_width

wb.save(file_path)

# Create team statistics sheet
for row in sheet.iter_rows(min_row=2, values_only=True): 
    cell_value = row[1]
    if cell_value in played_matches:
        played_matches[cell_value] += 1
    cell_value = row[2]
    if cell_value in played_matches:
        played_matches[cell_value] += 1
    cell_value = row[9]
    if cell_value == 1:
        if row[3] == "1" or row[3] == "1x" or row[3] == "lX" or row[3] == "lx" or row[3] == "IX" or row[3] == "Ix":
            won_matches[row[1]] += 1
        if row[3] == "2" or row[3] == "X2" or row[3] == "x2":
            won_matches[row[2]] += 1

sheet_name = "Teams"
sheet=wb[sheet_name]

sheet["A1"] = "Team"
sheet["B1"] = "Bet on"
sheet["C1"] = "Successful"
sheet["D1"] = "Success %"

team_index = 2
for k, v in leagues.items():
    for i in v:
        sheet[f"A{team_index}"] = i
        sheet[f"B{team_index}"] = played_matches[i]
        sheet[f"C{team_index}"] = won_matches[i]
        if played_matches[i] == 0:
             success_rate = 0
        else:
            success_rate = round(won_matches[i] / played_matches[i] * 100, 2)
        sheet[f"D{team_index}"] = str(success_rate)+ "%"
        team_index += 1

for row in range(1,sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row,col).font = Font(name='Helvetica', size=12, bold=True, color = '000000')
        sheet.cell(row,col).alignment = Alignment(horizontal='center', vertical='center')

for col in sheet.columns:
     max_width = 0
     column = col[0].column_letter
     for cell in col:
             if len(str(cell.value)) > max_width:
                 max_width = len(str(cell.value))
     set_col_width = max_width + 7
     sheet.column_dimensions[column].width = set_col_width

wb.save(file_path)

filelist = glob.glob(os.path.join(folder, "*.jpg"))
for f in filelist:
    os.remove(f)