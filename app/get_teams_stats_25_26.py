import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font

#https://www.whatismybrowser.com/detect/what-http-headers-is-my-browser-sending/
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
}

# dictionary with team fixtures
premierLeague = {
                "Arsenal" : "https://www.transfermarkt.com/fc-arsenal/spielplan/verein/11/saison_id/2025",
                "Chelsea" : "https://www.transfermarkt.com/fc-chelsea/spielplan/verein/631/saison_id/2025",
                "Liverpool" : "https://www.transfermarkt.com/fc-liverpool/spielplan/verein/31/saison_id/2025", 
                "ManchesterCity" : "https://www.transfermarkt.com/manchester-city/spielplan/verein/281/saison_id/2025",
                "ManchesterUnited" : "https://www.transfermarkt.com/manchester-united/spielplan/verein/985/saison_id/2025"
                }

laLiga = {
        "AtleticoMadrid" : "https://www.transfermarkt.com/atletico-madrid/spielplan/verein/13/saison_id/2025", 
        "Barcelona" : "https://www.transfermarkt.com/fc-barcelona/spielplan/verein/131/saison_id/2025",
        "RealMadrid" : "https://www.transfermarkt.com/real-madrid/spielplan/verein/418/saison_id/2025"
        }

bundesLiga = {
            "BayerLeverkusen" : "https://www.transfermarkt.com/bayer-04-leverkusen/spielplan/verein/15/saison_id/2025",
            "BayernMunchen" : "https://www.transfermarkt.com/fc-bayern-munchen/spielplan/verein/27/saison_id/2025", 
            "BorussiaDortmund" : "https://www.transfermarkt.com/borussia-dortmund/spielplan/verein/16/saison_id/2025", 
            "Frankfurt" : "https://www.transfermarkt.com/eintracht-frankfurt/spielplan/verein/24/saison_id/2025",
            "Mainz" : "https://www.transfermarkt.com/1-fsv-mainz-05/spielplan/verein/39/saison_id/2025",
            "RBLeipzig" : "https://www.transfermarkt.com/rasenballsport-leipzig/spielplan/verein/23826/saison_id/2025",
            "Stuttgart" : "https://www.transfermarkt.com/vfb-stuttgart/spielplan/verein/79/saison_id/2025"
            }

serieA = {
    "ACMilan" : "https://www.transfermarkt.com/ac-mailand/spielplan/verein/5/saison_id/2025", 
    "InterMilan" : "https://www.transfermarkt.com/inter-mailand/spielplan/verein/46/saison_id/2025",
    "Juventus" : "https://www.transfermarkt.com/juventus-turin/spielplan/verein/506/saison_id/2025",
    "Lazio" : "https://www.transfermarkt.com/lazio-rom/spielplan/verein/398/saison_id/2025",
    "Napoli" : "https://www.transfermarkt.com/ssc-neapel/spielplan/verein/6195/saison_id/2025", 
    "Roma" : "https://www.transfermarkt.com/as-rom/spielplan/verein/12/saison_id/2025"
        }   

ligue1 = {
        "Lille" : "https://www.transfermarkt.com/losc-lille/spielplan/verein/1082/saison_id/2025",
        "Lyon" : "https://www.transfermarkt.com/olympique-lyon/spielplan/verein/1041/saison_id/2025",
        "Marseille" : "https://www.transfermarkt.com/olympique-marseille/spielplan/verein/244/saison_id/2025",
        "Monaco" : "https://www.transfermarkt.com/as-monaco/spielplan/verein/162/saison_id/2025",
        "Nice" : "https://www.transfermarkt.com/ogc-nizza/spielplan/verein/417/saison_id/2025",
        "PSG" : "https://www.transfermarkt.com/fc-paris-saint-germain/spielplan/verein/583/saison_id/2025"
        }

leagues ={
        "England" : premierLeague,
        "Spain" : laLiga,
        "Germany" : bundesLiga,
        "Italy" : serieA,
        "France" : ligue1
        }

# generate new Excel sheet
file_path = "football_statistics.xlsx"
sheet_name = "Generated"
wb = Workbook()
wb.create_sheet(sheet_name)
wb.remove(wb["Sheet"])
sheet=wb[sheet_name]
sheet.freeze_panes = "C2"
wb.save(file_path)

# function to retrieve data from TransferMarkt
def retrieve_data(index, link, sect):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]
        
    response = requests.get(link, headers=headers)
    response.status_code
    soup = BeautifulSoup(response.content, "html.parser")

    section = soup.find('a', href=sect)
    table = section.find_parent('thead').find_next_sibling('tbody')
    data = [td.get_text(strip=True) for td in table.find_all('td', class_='zentriert')]

    for i, value in enumerate(data[:12]):
        if(value == '-'):
            data[i] = 0

    # compute each field
    homeMatches = int(data[0])
    homeWins = int(data[1])
    homeDraws = int(data[2])
    homeLosses = int(data[3])
    homePoints = round(float(data[4]), 2)

    if homeMatches == 0:
        homeWins_p = 0
        homeDraws_p = 0
        homeLosses_p = 0
        homeGoals_f = 0
        homeGoals_a = 0
        homeGoals_f_v = 0
        homeGoals_a_v = 0
    else:
        homeWins_p = round(homeWins / homeMatches * 100, 2)
        homeDraws_p = round(homeDraws / homeMatches * 100, 2)
        homeLosses_p = round(homeLosses / homeMatches* 100 , 2)
        homeGoals = data[5].split(":")
        homeGoals_f = int(homeGoals[0])
        homeGoals_a = int(homeGoals[1])
        homeGoals_f_v = round(homeGoals_f / homeMatches , 2)
        homeGoals_a_v = round(homeGoals_a / homeMatches, 2)
    homeGD = homeGoals_f - homeGoals_a

    awayMatches = int(data[6])
    awayWins = int(data[7])
    awayDraws = int(data[8])
    awayLosses = int(data[9])
    awayPoints = round(float(data[10]), 2)

    if awayMatches == 0:
        awayWins_p = 0
        awayDraws_p = 0
        awayLosses_p = 0
        awayGoals_f = 0
        awayGoals_a = 0
        awayGoals_f_v = 0
        awayGoals_a_v = 0
    else:
        awayWins_p = round(awayWins / awayMatches * 100, 2)
        awayDraws_p = round(awayDraws / awayMatches * 100, 2)
        awayLosses_p = round(awayLosses / awayMatches* 100 , 2)
        awayGoals = data[11].split(":")
        awayGoals_f = int(awayGoals[0])
        awayGoals_a = int(awayGoals[1])
        awayGoals_f_v = round(awayGoals_f / awayMatches, 2)
        awayGoals_a_v = round(awayGoals_a / awayMatches, 2)
    awayGD = awayGoals_f - awayGoals_a

    totalMatches = homeMatches + awayMatches
    totalWins = homeWins + awayWins
    totalDraws = homeDraws + awayDraws
    totalLosses = homeLosses + awayLosses
    totalGoals_f = homeGoals_f + awayGoals_f
    totalGoals_a = homeGoals_a + awayGoals_a
    if totalMatches == 0:
        totalWins_p = 0
        totalDraws_p = 0
        totalLosses_p = 0
        totalPoints = 0
        totalGoals_f_v = 0
        totalGoals_a_v = 0
    else:
        totalWins_p = round(totalWins / totalMatches * 100, 2)
        totalDraws_p = round(totalDraws / totalMatches * 100 , 2)
        totalLosses_p = round(totalLosses / totalMatches * 100, 2)
        totalPoints = round((3 * totalWins + totalDraws) / totalMatches, 2)
        totalGoals_f_v = round(totalGoals_f / totalMatches, 2)
        totalGoals_a_v = round(totalGoals_a / totalMatches, 2)
    totalGD = totalGoals_f - totalGoals_a

    #update Excel
    sheet[f"C{index}"] = homeMatches
    sheet[f"D{index}"] = homeWins
    sheet[f"E{index}"] = homeWins_p
    sheet[f"F{index}"] = homeDraws
    sheet[f"G{index}"] = homeDraws_p
    sheet[f"H{index}"] = homeLosses
    sheet[f"I{index}"] = homeLosses_p
    sheet[f"J{index}"] = homePoints
    sheet[f"K{index}"] = homeGoals_f
    sheet[f"L{index}"] = homeGoals_f_v
    sheet[f"M{index}"] = homeGoals_a
    sheet[f"N{index}"] = homeGoals_a_v
    sheet[f"O{index}"] = homeGD

    sheet[f"P{index}"] = awayMatches
    sheet[f"Q{index}"] = awayWins
    sheet[f"R{index}"] = awayWins_p
    sheet[f"S{index}"] = awayDraws
    sheet[f"T{index}"] = awayDraws_p
    sheet[f"U{index}"] = awayLosses
    sheet[f"V{index}"] = awayLosses_p
    sheet[f"W{index}"] = awayPoints
    sheet[f"X{index}"] = awayGoals_f
    sheet[f"Y{index}"] = awayGoals_f_v
    sheet[f"Z{index}"] = awayGoals_a
    sheet[f"AA{index}"] = awayGoals_a_v
    sheet[f"AB{index}"] = awayGD

    sheet[f"AC{index}"] = totalMatches
    sheet[f"AD{index}"] = totalWins
    sheet[f"AE{index}"] = totalWins_p
    sheet[f"AF{index}"] = totalDraws
    sheet[f"AG{index}"] = totalDraws_p
    sheet[f"AH{index}"] = totalLosses
    sheet[f"AI{index}"] = totalLosses_p
    sheet[f"AJ{index}"] = totalPoints
    sheet[f"AK{index}"] = totalGoals_f
    sheet[f"AL{index}"] = totalGoals_f_v
    sheet[f"AM{index}"] = totalGoals_a
    sheet[f"AN{index}"] = totalGoals_a_v
    sheet[f"AO{index}"] = totalGD

    wb.save(file_path)

# call function based on league
index = 2
sect = ""

for k,v in leagues.items():
    for i in v:
        if k == "England":
            sect = "/premier-league/startseite/wettbewerb/GB1/saison_id/2025"
        elif k == "Spain":
            sect = "/laliga/startseite/wettbewerb/ES1/saison_id/2025"
            if index == 2 + len(premierLeague):
                index +=2
        elif k == "Germany":
            sect = "/bundesliga/startseite/wettbewerb/L1/saison_id/2025"
            if index == 2 + len(premierLeague) + 2 + len(laLiga):
                index += 2
        elif k == "Italy":
            sect = "/serie-a/startseite/wettbewerb/IT1/saison_id/2025"
            if index == 2 + len(premierLeague) + 2 + len(laLiga) + 2 + len(bundesLiga):
                index += 2
        elif k == "France":
            sect = "/ligue-1/startseite/wettbewerb/FR1/saison_id/2025"
            if index == 2 + len(premierLeague) + 2 + len(laLiga) + 2 + len(bundesLiga) + 2 + len(serieA):
                index += 2
        retrieve_data(index, v[i], sect)
        index +=1

# populate Excel sheet with team names and metrics
wb = load_workbook(file_path)
sheet = wb[sheet_name]

for row in range(1,sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        sheet.cell(row,col).font = Font(name='Helvetica', size=12, bold=True, color = '000000')
        sheet.cell(row,col).alignment = Alignment(horizontal='center', vertical='center')
        

sheet["A1"] = "Premier League"
sheet["B1"] = "Season"
sheet["C1"] = "Home matches"
sheet["D1"] = "Home wins"
sheet["E1"] = "%"
sheet["F1"] = "Home draws"
sheet["G1"] = "%"
sheet["H1"] = "Home losses"
sheet["I1"] = "%"
sheet["J1"] = "Home points / match"
sheet["K1"] = "Home goals scored"
sheet["L1"] = " / match"
sheet["M1"] = "Home goals conceded"
sheet["N1"] = " / match"
sheet["O1"] = "Home goal difference"
sheet["P1"] = "Away matches"
sheet["Q1"] = "Away wins"
sheet["R1"] = "%"
sheet["S1"] = "Away draws"
sheet["T1"] = "%"
sheet["U1"] = "Aways losses"
sheet["V1"] = "%"
sheet["W1"] = "Aways points / match"
sheet["X1"] = "Aways goals scored"
sheet["Y1"] = " / match"
sheet["Z1"] = "Away goals conceded"
sheet["AA1"] = " / match"
sheet["AB1"] = "Away home difference"
sheet["AC1"] = "Total matches"
sheet["AD1"] = "Total wins"
sheet["AE1"] = "%"
sheet["AF1"] = "Total draws"
sheet["AG1"] = "%"
sheet["AH1"] = "Total losses"
sheet["AI1"] = "%"
sheet["AJ1"] = "Total points / match"
sheet["AK1"] = "Total goals scored"
sheet["AL1"] = " / match"
sheet["AM1"] = "Total goals conceded"
sheet["AN1"] = " / match"
sheet["AO1"] = "Total goal difference"

index = 2

for k,v in leagues.items():
    for i in v:
        if k == "Spain":
            if index == 2 + len(premierLeague):
                sheet[f"A{index+1}"] = "La Liga"
                index +=2
        elif k == "Germany":
            if index == 2 + len(premierLeague) + 2 + len(laLiga):
                sheet[f"A{index+1}"] = "Bundesliga"
                index += 2
        elif k == "Italy":
            if index == 2 + len(premierLeague) + 2 + len(laLiga) + 2 + len(bundesLiga):
                sheet[f"A{index+1}"] = "Serie A"
                index += 2
        elif k == "France":
            if index == 2 + len(premierLeague) + 2 + len(laLiga) + 2 + len(bundesLiga) + 2 + len(serieA):
                sheet[f"A{index+1}"] = "Ligue 1"
                index += 2
        sheet[f"A{index}"] = i
        sheet[f"B{index}"] = "2025/2026"
        index +=1

for col in sheet.columns:
     max_width = 0
     column = col[0].column_letter
     for cell in col:
             if len(str(cell.value)) > max_width:
                 max_width = len(str(cell.value))
     set_col_width = max_width + 7
     sheet.column_dimensions[column].width = set_col_width
     
wb.save(file_path)